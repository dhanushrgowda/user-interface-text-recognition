import openpyxl
import pytesseract as tes
from PIL import Image




def openandwrite(fname,image_list,column_names=None):
    print("entered open and write")
    list_of_images=image_list
    global wb
    global s
    excel_name=fname
    wb=openpyxl.load_workbook(fname)
    s=wb.get_sheet_by_name("Sheet1")
    s.column_dimensions['A'].width=20
    s.column_dimensions['B'].width=20
    s.column_dimensions['C'].width=20
    s.column_dimensions['D'].width=35
    s.column_dimensions['E'].width=50
    if column_names != None:
        lst=column_names.split(" ")
        for i in range(0,len(lst)):
            s.cell(row=1,column=(i+1)).value=(lst[i].upper())
            print("written one row")
        image_to_file(list_of_images,excel_name)
    else:
        print("writing direct")
        image_to_file(list_of_images,excel_name)
    wb.save(fname)
def image_to_file(image_list,excel_name):
    mat=[[0 for x in range(5) ]for y in range(len(image_list))]
    print(mat)
    for i in range(len(image_list)):
        img=Image.open(image_list[i])
        text=tes.image_to_string(img,lang='eng')
        details=[]
        details=text.split('\n')
        col=0
        print('i=',i)
        for j in details:
            if len(j)>3:
                print('col=',col)
                mat[i][col]=j
                col+=1
    print("mat 2D list")
    print(mat)
    for row in mat:
        r=2
        c=1
        for j in row:
            s.cell(row=r,column=c).value=j
            c+=1
        r+=1
    print("finished writing")
    excel_name=excel_name+'.xlsx'
    #wb.save(excel_name)
    
    
          
        
